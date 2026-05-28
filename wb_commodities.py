"""
Збирач місячних цін сировини зі Світового банку (Pink Sheet) → PostgreSQL.

Джерело (безкоштовно, без ключа): CMO-Historical-Data-Monthly.xlsx
  аркуш 'Monthly Prices': рядок 5 — назви, рядок 6 — одиниці, з рядка 7 — дані;
  стовпець A — місяць 'РРРРMмм' (напр. '2026M04'); '…' = пропуск.

Збираємо релевантне для ціноутворення е/е в Україні:
  Natural gas, Europe ($/mmbtu)  — TTF-проксі (паливо граничних блоків)
  Coal, Australian / South African ($/mt)
  Crude oil, Brent ($/bbl)
  (+ Natural gas, US для контексту)

Використання:
  python wb_commodities.py                       # з 2020-01, поточний URL
  python wb_commodities.py --start 2015-01-01
  python wb_commodities.py --url <інший_лінк>     # якщо URL зміниться
  python wb_commodities.py --file /tmp/wb.xlsx    # з локального файлу

Середовище: OREE_DSN
Залежності: httpx, asyncpg, openpyxl

Примітка: лінк WB містить ідентифікатор видання і час від часу змінюється.
Якщо завантаження дає не-xlsx — оновіть URL зі сторінки
worldbank.org/en/research/commodity-markets (--url).
"""
from __future__ import annotations

import argparse
import asyncio
import io
import logging
import os
import re
import sys
from datetime import date

import asyncpg
import httpx
import openpyxl

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(name)s | %(message)s")
log = logging.getLogger("oree.wb")

DEFAULT_URL = ("https://thedocs.worldbank.org/en/doc/"
               "18675f1d1639c7a34d463f59263ba0a2-0050012025/related/"
               "CMO-Historical-Data-Monthly.xlsx")
UA = "oree-research-collector/0.1 (market research)"
SHEET = "Monthly Prices"
DATE_RE = re.compile(r"^(\d{4})M(\d{2})$")

# чистий код -> початок назви колонки у файлі WB
WANT = {
    "oil_brent":      "Crude oil, Brent",
    "gas_europe":     "Natural gas, Europe",
    "gas_us":         "Natural gas, US",
    "coal_australia": "Coal, Australian",
    "coal_safrica":   "Coal, South African",
}

DDL = """
CREATE TABLE IF NOT EXISTS commodity_prices (
    month       DATE         NOT NULL,
    commodity   VARCHAR(40)  NOT NULL,
    unit        VARCHAR(20),
    value       NUMERIC(14,4),
    source      VARCHAR(20)  NOT NULL DEFAULT 'worldbank',
    ingested_at TIMESTAMPTZ  NOT NULL DEFAULT now(),
    PRIMARY KEY (month, commodity)
);
"""

UPSERT = """
INSERT INTO commodity_prices (month, commodity, unit, value)
VALUES ($1, $2, $3, $4)
ON CONFLICT (month, commodity) DO UPDATE
SET unit = EXCLUDED.unit, value = EXCLUDED.value, ingested_at = now();
"""


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "…", "..", "...", "n.a.", "n/a", "-"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def parse_rows(rows, start=None):
    """rows: ітерабельний список кортежів (як iter_rows values_only)."""
    rows = [list(r) for r in rows]
    hdr = None
    for i, r in enumerate(rows):
        if any(isinstance(c, str) and c.strip().startswith("Crude oil, Brent") for c in r):
            hdr = i
            break
    if hdr is None:
        return []
    header = rows[hdr]
    units = rows[hdr + 1] if hdr + 1 < len(rows) else [None] * len(header)

    colmap = {}  # clean -> (col_idx, unit)
    for clean, sub in WANT.items():
        for ci, c in enumerate(header):
            if isinstance(c, str) and c.strip().startswith(sub):
                u = units[ci] if ci < len(units) else None
                u = (str(u).strip("() ") if u else None) or None
                colmap[clean] = (ci, u)
                break

    out = []
    for r in rows[hdr + 2:]:
        if not r or r[0] is None:
            continue
        m = DATE_RE.match(str(r[0]).strip())
        if not m:
            continue
        dt = date(int(m.group(1)), int(m.group(2)), 1)
        if start and dt < start:
            continue
        for clean, (ci, unit) in colmap.items():
            val = _num(r[ci]) if ci < len(r) else None
            if val is None:
                continue
            out.append((dt, clean, unit, val))
    return out


def load_workbook_bytes(url=None, file=None):
    if file:
        with open(file, "rb") as f:
            data = f.read()
    else:
        r = httpx.get(url, timeout=120, follow_redirects=True,
                      headers={"User-Agent": UA})
        r.raise_for_status()
        data = r.content
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)


async def run(start, url, file, dsn):
    wb = load_workbook_bytes(url=url, file=file)
    name = SHEET if SHEET in wb.sheetnames else next(
        (s for s in wb.sheetnames if "onthly" in s.lower() and "price" in s.lower()),
        wb.sheetnames[0])
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    parsed = parse_rows(rows, start=start)
    if not parsed:
        log.error("нічого не розпарсено (перевірте аркуш/URL)")
        return 1

    pool = await asyncpg.create_pool(dsn=dsn, min_size=1, max_size=2)
    try:
        async with pool.acquire() as c:
            await c.execute(DDL)
            async with c.transaction():
                await c.executemany(UPSERT, parsed)
        by_c = {}
        for dt, clean, _u, _v in parsed:
            by_c.setdefault(clean, []).append(dt)
        for clean, dts in sorted(by_c.items()):
            log.info("[%s] %d міс. (%s → %s)", clean, len(dts), min(dts), max(dts))
        log.info("готово. рядків=%d", len(parsed))
        return 0
    finally:
        await pool.close()


def main():
    ap = argparse.ArgumentParser(description="World Bank Pink Sheet commodity collector")
    ap.add_argument("--start", default="2020-01-01", help="YYYY-MM-DD (фільтр з)")
    ap.add_argument("--url", default=DEFAULT_URL)
    ap.add_argument("--file", help="локальний xlsx замість завантаження")
    ap.add_argument("--dsn", default=os.environ.get(
        "OREE_DSN", "postgresql://oree:postgres@localhost:5432/oree"))
    a = ap.parse_args()
    start = date.fromisoformat(a.start)
    return asyncio.run(run(start, a.url, a.file, a.dsn))


if __name__ == "__main__":
    sys.exit(main())
